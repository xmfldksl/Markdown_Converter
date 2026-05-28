import os # 파일의 확장자를 분리하고 경로를 분석하기 위한 파이썬 내장 모듈을 가져옵니다.
import pandas as pd # 1만 줄 단위로 분할된 데이터를 데이터프레임으로 묶어 마크다운 표로 한 번에 변환하기 위한 외부 라이브러리를 가져옵니다.
import openpyxl # 표준 XML 기반의 엑셀 파일(.xlsx)을 스트리밍 방식으로 읽어오기 위한 라이브러리를 가져옵니다.
from pyxlsb import open_workbook as open_xlsb # 바이너리 기반의 대용량 엑셀 파일(.xlsb)을 스트리밍 방식으로 읽어오기 위한 라이브러리 함수를 지정하여 가져옵니다.

def extract_excel_content(file_path, progress_callback=None):
    # 파일 경로를 받아 확장자에 따라 적절한 엑셀 파서를 선택하여 마크다운 텍스트로 반환하는 메인 추출 함수입니다.
    md_output = "" # 최종 결과물이 누적될 빈 텍스트 문자열 변수를 초기화합니다.
    
    _, ext = os.path.splitext(file_path) # 파일 경로에서 마침표를 기준으로 파일 이름과 확장자를 분리합니다.
    ext_lower = ext.lower() # 확장자를 모두 소문자 문자열로 강제 변환합니다.

    if ext_lower == '.xlsb': # 만약 입력된 파일의 확장자가 바이너리 엑셀(.xlsb) 파일이라면.
        with open_xlsb(file_path) as wb: # pyxlsb 라이브러리를 사용하여 바이너리 엑셀 파일을 메모리에 엽니다.
            sheet_names = wb.sheets # 파일 내부에 존재하는 모든 워크시트의 이름들을 리스트 형태로 가져옵니다.
            total_sheets = len(sheet_names) # 전체 시트의 개수를 세어 퍼센트 진행률 계산의 기준값으로 삼습니다.
            
            for i, sheet_name in enumerate(sheet_names): # 첫 번째 시트부터 순서대로 번호표를 붙여가며 반복 탐색합니다.
                md_output += f"## Sheet: {sheet_name}\n\n" # 시트의 이름을 마크다운 중제목 문법으로 조립하여 결과물에 추가합니다.
                
                with wb.get_sheet(sheet_name) as sheet: # 현재 시트의 이름으로 접근하여 시트 객체를 전용 메모리 공간에 엽니다.
                    rows = sheet.rows() # 모든 행 데이터를 차례대로 뱉어내는 반복자 객체를 생성합니다.
                    
                    try: # 빈 시트를 읽었을 때 프로그램이 폭발하는 것을 막기 위해 예외 처리 구문으로 감쌉니다.
                        first_row = next(rows) # 첫 번째 줄 데이터를 하나 뽑아내어 제목 행 후보로 삼습니다.
                    except StopIteration: # 시트에 데이터가 없어서 뽑아낼 것이 없는 예외가 발생한다면.
                        continue # 아래의 텍스트 추출 과정을 생략하고 바로 다음 시트 탐색으로 건너뜁니다.
                    
                    # 제목 행의 각 칸이 비어있지 않다면 문자열로 바꾸고 마크다운 표를 부수는 줄바꿈과 파이프 기호를 이스케이프 처리하여 리스트로 만듭니다.
                    headers = [str(cell.v).replace('\n', '<br>').replace('|', '\\|') if cell.v is not None else "" for cell in first_row]
                    
                    if not any(headers): # 만약 첫 번째 줄이 완전히 텅 비어있다면.
                        headers = [f"Column_{j}" for j in range(len(headers))] # 칸의 개수만큼 임시 제목 문자열을 강제로 생성합니다.
                    
                    chunk = [] # 1만 줄 단위로 데이터를 모아두기 위한 빈 리스트 공간을 생성합니다.
                    for row in rows: # 두 번째 줄부터 마지막 줄까지 순서대로 반복 탐색합니다.
                        # 각 데이터 칸이 비어있지 않다면 문자열로 바꾸고 마크다운 충돌 기호들을 치환 처리하여 리스트로 묶습니다.
                        clean_row = [str(cell.v).replace('\n', '<br>').replace('|', '\\|') if cell.v is not None else "" for cell in row]
                        chunk.append(clean_row) # 정제된 한 줄의 데이터를 1만 줄 모음통 리스트에 추가합니다.
                        
                        if len(chunk) >= 10000: # 모음통에 데이터가 1만 줄 이상 꽉 찼는지 검사합니다.
                            df = pd.DataFrame(chunk, columns=headers) # 2차원 판다스 데이터프레임 구조로 찍어냅니다.
                            md_output += df.to_markdown(index=False) + "\n\n" # 데이터프레임 전체를 마크다운 표 텍스트로 일괄 변환하여 결과물에 이어 붙입니다.
                            chunk = [] # 모음통 리스트를 텅 비웁니다.
                            
                    if chunk: # 반복문이 끝난 뒤 남아있는 나머지 짜투리 데이터가 존재하는지 검사합니다.
                        df = pd.DataFrame(chunk, columns=headers) # 짜투리 데이터를 판다스 데이터프레임으로 변환합니다.
                        md_output += df.to_markdown(index=False) + "\n\n" # 나머지 데이터프레임을 마크다운으로 변환하여 최종 결과물에 누적시킵니다.
                
                if progress_callback: # UI 스레드로 진행률을 보고할 함수가 연결되어 있는지 확인합니다.
                    progress_callback(int(((i + 1) / total_sheets) * 100)) # 정수 형태의 백분율 수치만 깔끔하게 UI로 전송합니다.

    else: # 입력된 파일의 확장자가 바이너리가 아닌 일반 엑셀(.xlsx) 파일이라면.
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True) # 대용량 처리를 위한 읽기 전용 모드와 계산값 도출 모드를 켜서 엑셀 파일을 엽니다.
        
        sheet_names = wb.sheetnames # 엑셀 파일 내부에 기록된 워크시트 이름 리스트를 가져옵니다.
        total_sheets = len(sheet_names) # 퍼센트 계산을 위해 시트 총개수를 파악합니다.
        
        for i, sheet_name in enumerate(sheet_names): # 전체 시트를 하나씩 반복 탐색합니다.
            sheet = wb[sheet_name] # 시트 이름으로 데이터 객체를 메모리에 호출합니다.
            md_output += f"## Sheet: {sheet_name}\n\n" # 마크다운 중제목 형태로 시트 이름을 표기합니다.
            
            rows = sheet.iter_rows(values_only=True) # 텍스트 값만 뽑아내는 초고속 반복자를 생성합니다.
            
            try: # 빈 시트 처리 에러 방어 로직입니다.
                first_row = next(rows) # 첫 번째 행을 제목 줄로 지정합니다.
            except StopIteration: # 뽑아낼 데이터가 없다면.
                continue # 건너뜁니다.
            
            headers = [str(h).replace('\n', '<br>').replace('|', '\\|') if h is not None else "" for h in first_row] 
            
            if not any(headers): # 제목 줄이 비어있다면.
                headers = [f"Column_{j}" for j in range(len(headers))] # 임시 컬럼명을 생성합니다.
                
            chunk = [] # 1만 줄 분할 처리용 빈 리스트입니다.
            for row in rows: # 남은 데이터 행들을 순회합니다.
                clean_row = [str(cell).replace('\n', '<br>').replace('|', '\\|') if cell is not None else "" for cell in row] 
                chunk.append(clean_row) # 리스트에 추가합니다.
                
                if len(chunk) >= 10000: # 1만 줄 단위로 끊어냅니다.
                    df = pd.DataFrame(chunk, columns=headers) # 판다스 객체로 만듭니다.
                    md_output += df.to_markdown(index=False) + "\n\n" # 표 텍스트로 변환합니다.
                    chunk = [] # 메모리를 비웁니다.
                    
            if chunk: # 남은 데이터가 있다면.
                df = pd.DataFrame(chunk, columns=headers) # 데이터프레임으로 감쌉니다.
                md_output += df.to_markdown(index=False) + "\n\n" # 텍스트 변환하여 붙여넣습니다.
                
            if progress_callback: # 해당 시트의 변환이 완전히 종료되었다면.
                percent = int(((i + 1) / total_sheets) * 100) # 진행률을 계산합니다.
                progress_callback(percent) # 메인 스레드로 순수 정수 숫자만 전송합니다.
                
        wb.close() # openpyxl 파일 객체를 명시적으로 닫습니다.

    return md_output # 분석과 조립이 모두 끝난 거대한 마크다운 문자열을 메인 프로그램으로 반환합니다.